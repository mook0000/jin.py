# int_code = {}
# in_file =open("C:/pythonfiles/lims/files/interest.txt", "r")
#
# for line in in_file:
#     code,c_name = line.split()
#     # code = line[0:6]
#     # c_name = line[7:]
#     int_code.update({code:{'name':c_name,'시험':1}})
#
# print(int_code)
# for codelist in int_code:
#     print(codelist)
temp =None
lambda temp: 0 if temp==None else temp
print(temp,type(temp))
x=90
check_score = lambda x: 'pass' if x>=60 else 'fail'
print(check_score)

test_list = ['lim','0.5','53 ']
for ui in test_list:
    if '.' in ui :
        x=float(ui)
        print(x)
        continue
    try:
        x=int(ui)
        print(x)
    except:
        x=ui.split()
    print(x)
    #     print(data1)
    # c_name=pickle.load(file)
    # for line in file:
    #     a=[]
    #     a=line.strip(" ")
    #     print(a)

from openpyxl import Workbook
# from openpyxl import load_workbook
#
# wb = load_workbook('./Test.xlsm',keep_vba=True)
# ws = wb['Sheet1']
# ws.cell(row=1,column=1).value = 9999999
# wb.save('./Test.xlsm')
a = [1,2,3,2,45,2,5]
enumerate(a)
print(a)
print(list(enumerate(a)))
i=0
for v in a:
    print(i,v)
    i += 1
for a,b in enumerate(a):
    print(f'{a+1};{b}')